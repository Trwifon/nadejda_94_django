from django.db.models import Sum
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font

MIN_AREA = 0.3
MAX_AREA = 4.0
CORRECTION = 1.5

def calculate_area(width, height, number):

    area = round(width * height / 1000000, 2)

    if area < MIN_AREA:
        area = MIN_AREA
    elif area > MAX_AREA:
        area *= CORRECTION

    area *= number

    return area


def calculate_price(area, unit_price, supplement):
    price = round((unit_price * area),2) + int(supplement)

    return price


def calculate_glass_data(ALL_ORDERS):
    glass_data = {
        'total_number': 0,
        'total_area': 0,
        'total_price': 0,
    }

    for order in ALL_ORDERS:
        area = calculate_area(order['width'], order['height'], order['number'])
        supplement = order['supplement'] if order['supplement'] else 0
        order['price'] = calculate_price(area, float(order['unit_price']), supplement)

        glass_data['total_number'] += order['number']
        glass_data['total_area'] += area

        glass_data['total_price'] += order['price']

    return glass_data


def get_glass_kind(order):
    if order.third_glass:
        kind = (f"{order.first_glass}"
                f"+{order.second_glass}"
                f"+{order.third_glass}"
                f"={order.thickness} ")

    elif order.second_glass:
        kind = (f"{order.first_glass}"
                f"+{order.second_glass}"
                f"={order.thickness} ")

    else:
        kind = (f"{order.first_glass} ")

    if order.module:
        kind += f"{order.module}"

    return kind


def get_additional_fields(glass, price_with_accumulation):
    simple_area = calculate_area(glass.width, glass.height, glass.number)
    price_for_all_units = simple_area * float(glass.unit_price)

    total_price_for_glass = price_for_all_units + glass.supplement

    price_with_accumulation += total_price_for_glass

    additional_fields = [
        glass.id,
        glass.unit_price,
        simple_area,
        price_for_all_units,
        glass.supplement,
        total_price_for_glass,
        price_with_accumulation
    ]

    return additional_fields


def excel_glass_view(glasses):
    number_of_glasses = glasses.aggregate(sum=Sum('number'))
    total_area = 0.0
    glass_order = []
    glass_order_with_additional_fields = []
    dist_order = []
    row = 0
    old_order = ''
    order_area = 0
    price_with_accumulation = 0

    for glass in glasses:
        current_order = glass.record.order
        quantity = glasses.filter(record__order=current_order).aggregate(sum=Sum('number'))
        if current_order == old_order:
            row += 1
        else:
            row = 1
            order_area = 0
            record_glasses = glasses.filter(record__order=current_order).order_by('pk')

            for record_glass in record_glasses:
                area = calculate_area(record_glass.width, record_glass.height, record_glass.number)
                order_area += area

        old_order = current_order

        if glass.record.note == 'None' or not glass.record.note:
            first_column = f"{glass.record.partner.name} / {quantity['sum']}"
        else:
            if glass.record.partner.name == 'Клиент':
                first_column = f"{glass.record.note} / {quantity['sum']}"
            else:
                first_column = f"{glass.record.partner.name} / {quantity['sum']}"

        glass_order_new_row = [
            first_column,
            f"{current_order} {row}",
            glass.width,
            glass.height,
            'R',
            glass.number,
            glass.number,
            get_glass_kind(glass),
        ]
        glass_order.append(glass_order_new_row)

        additional_fields = get_additional_fields(glass, price_with_accumulation)
        price_with_accumulation = additional_fields[-1]

        glass_order_with_additional_fields.append(glass_order_new_row + additional_fields)

        if row == 1:
            glass_order[-1].extend([order_area])

        dist_row = f"{glass.width},{glass.height},5,{glass.number}"
        dist_order.append(dist_row)

        total_area += calculate_area(glass.width, glass.height, glass.number)

    glass_order[0].extend([
        'Брой:',
        number_of_glasses['sum'],
        'Площ:',
        total_area,
        'кв.м'
    ])

    excel_view = {
        'glass_order_list': glass_order,
        'glass_order_with_additional_fields': glass_order_with_additional_fields,
        'dist_order_list': dist_order,
    }

    return excel_view


def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 4
    ws.column_dimensions['G'].width = 4
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 5

    for cell in ws['C']:
        cell.font = Font(bold=True)
    for cell in ws['D']:
        cell.font = Font(bold=True)
    for cell in ws['I']:
        cell.font = Font(bold=True)

    for cell in ws['C']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['D']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['E']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['F']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['G']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['H']:
        cell.alignment = Alignment(horizontal='left')

    ws['J1'].alignment = Alignment(horizontal='right')
    ws['K1'].alignment = Alignment(horizontal='left')
    ws['L1'].alignment = Alignment(horizontal='right')
    ws['M1'].alignment = Alignment(horizontal='left')

    for row in range(1, ws.max_row + 1):
        ws[f'G{row}'] = f'=F{row}'

    ws.sheet_view.zoomScale = 130

    wb.save(file_path)

    return None
